"""KOSDAQ 우선주 자사주 전수 + 종류주식 네이밍 변종 수집(혼재 점검)."""
import warnings as W; W.filterwarnings("ignore")
import asyncio, os, sys, io, re, json, time
from collections import Counter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)
from pathlib import Path

from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parent.parent / ".env")  # OPENDART_API_KEY 등 (repo 표준)
from openpyxl import load_workbook
from open_proxy_mcp.services.company import resolve_company_query
from open_proxy_mcp.services.filing_search import search_filings_by_report_name
from open_proxy_mcp.dart.client import get_dart_client
from open_proxy_mcp.services.treasury_share import _acode_int
c=get_dart_client()
def kq_top(n):
    wb=load_workbook("sample_universe/general_universe.xlsx",read_only=True,data_only=True)
    ws=wb[wb.sheetnames[0]]; recs=[]
    for r in ws.iter_rows(min_row=14,values_only=True):
        code,name,_,mcap,mkt=r[0],r[1],r[2],r[3],r[4]
        if code and name and mkt=="KQ" and mcap not in (None,""):
            recs.append((str(name).strip(),float(mcap)))
    recs.sort(key=lambda x:-x[1]); return [x[0] for x in recs[:n]]
DAILY=re.compile(r"([\d,]{4,})\s+[가-힣A-Za-z()·.&\s]{2,30}?\s+\d{8}(?![\d-])")
KIND=re.compile(r"상환전환우선주|전환우선주|상환우선주|기타주식|종류주식|우선주식|우선주|보통주식|RCPS|제?\d+우[A-Z]?")
HDR=re.compile(r"우선주\s*\([^)]{1,30}\)")
def names_in(flat):
    return Counter(KIND.findall(flat)), HDR.findall(flat)
async def scan(cc,kw,acode):
    items,_,_=await search_filings_by_report_name(corp_code=cc,bgn_de="20230101",end_de="20260620",
        pblntf_tys="",keywords=[kw],max_pages=4)
    rows=[]; allkinds=Counter(); hdrs=set()
    for it in items[:6]:
        html=(await c.get_document_cached(it["rcept_no"])).get("html","") or ""
        flat=re.sub(r"\s+"," ",re.sub(r"<[^>]+>"," ",html))
        amt=_acode_int(html,acode); ds=sum(int(a.replace(",",""))for a in DAILY.findall(flat))
        pref=bool(re.search(r"우선주|기타주식|종류주식|RCPS",flat))
        kc,hd=names_in(flat); allkinds+=kc; hdrs|=set(hd)
        uc = ds>=1e8 and (amt is None or (ds>(amt or 0)*1.05 and ds-(amt or 0)>=1e8)) and pref
        rows.append({"dt":it["rcept_dt"],"acode":amt,"daily":ds,"undercount":uc})
    return rows,allkinds,hdrs
async def one(name):
    try:
        r=await resolve_company_query(name)
        if not r.selected: return {"name":name,"err":"resolve"}
        cc=r.selected["corp_code"]
        a,ak,ah=await scan(cc,"자기주식취득결과","ACQ_AMT")
        d,dk,dh=await scan(cc,"자기주식처분결과","DSP_AMT")
        return {"name":name,"acq":a,"dsp":d,"kinds":dict(ak+dk),"hdrs":sorted(ah|dh)}
    except Exception as e: return {"name":name,"err":f"{type(e).__name__}"}
async def main():
    n=int(sys.argv[1]) if len(sys.argv)>1 else 200
    names=kq_top(n); print(f"KOSDAQ 상위 {len(names)}사")
    res,t0=[],time.monotonic()
    for i in range(0,len(names),16):
        res.extend(await asyncio.gather(*(one(x) for x in names[i:i+16])))
        d=min(i+16,len(names))
        if d%48==0 or d==len(names): print(f"  ... {d}/{len(names)} ({time.monotonic()-t0:.0f}s)")
    pref=[r for r in res if r.get("kinds") and any(k!='보통주식' for k in r["kinds"])]
    acq_uc=[(r["name"],e) for r in res if r.get("acq") for e in r["acq"] if e["undercount"]]
    dsp_uc=[(r["name"],e) for r in res if r.get("dsp") for e in r["dsp"] if e["undercount"]]
    allkinds=Counter()
    for r in res:
        for k,v in (r.get("kinds") or {}).items(): allkinds[k]+=v
    print(f"\n종류주식 활동 기업: {len(pref)}사")
    print(f"\n=== 종류주식 네이밍 변종 빈도(전 KOSDAQ result/처분 본문) ===")
    for k,v in allkinds.most_common(): print(f"  {k}: {v}")
    print(f"\n=== 우선주 헤더명 샘플 ===")
    hdrs=set()
    for r in res: hdrs|=set(r.get("hdrs") or [])
    for h in sorted(hdrs)[:25]: print(f"  {h}")
    print(f"\n=== 취득결과 undercount: {len(acq_uc)} / 처분결과 undercount: {len(dsp_uc)} ===")
    for nm,e in acq_uc: print(f"  취득 {nm} {e['dt']}: ACODE {(e['acode'] or 0)/1e8:.0f}억 → 일별 {e['daily']/1e8:.0f}억")
    for nm,e in dsp_uc: print(f"  처분 {nm} {e['dt']}: ACODE {(e['acode'] or 0)/1e8:.0f}억 → 일별 {e['daily']/1e8:.0f}억")
    json.dump(res,open("wiki/architecture/audits/data/260617_treasury_kosdaq_naming.json","w",encoding="utf-8"),ensure_ascii=False,indent=2)
    print("\n저장: 260617_treasury_kosdaq_naming.json")
asyncio.run(main())
