"""우선주 보유 주요기업 자사주 공시 전수 — 취득/처분 결과의 복수종류(보통주+우선주) 누락 점검.
각 result 본문에서 ACODE 단일값 vs 일별 취득가액총액 합산 비교 → 누락(undercount) 탐지.
usage: uv run python scripts/treasury_multitype_census.py [N]
"""
import warnings as W; W.filterwarnings("ignore")
import asyncio, os, sys, io, re, json, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)
for line in open(".env.local", encoding="utf-8"):
    if line.startswith("DART_API_KEY="):
        os.environ["OPENDART_API_KEY"] = line.split("=",1)[1].strip()
from openpyxl import load_workbook
from open_proxy_mcp.services.company import resolve_company_query
from open_proxy_mcp.services.filing_search import search_filings_by_report_name
from open_proxy_mcp.dart.client import get_dart_client
from open_proxy_mcp.services.treasury_share import _acode_int
c=get_dart_client()

def ks_top(n):
    wb=load_workbook("sample_universe/general_universe.xlsx",read_only=True,data_only=True)
    ws=wb[wb.sheetnames[0]]; recs=[]
    for r in ws.iter_rows(min_row=14,values_only=True):
        code,name,_,mcap,mkt=r[0],r[1],r[2],r[3],r[4]
        if code and name and mkt=="KS" and mcap not in (None,""):
            recs.append((str(name).strip(),float(mcap)))
    recs.sort(key=lambda x:-x[1]); return [x[0] for x in recs[:n]]

DAILY=re.compile(r"([\d,]{4,})\s+[가-힣A-Za-z()·.&\s]{2,30}?\s+\d{8}(?![\d-])")
async def parse_results(cc, kw, acode):
    items,_,_=await search_filings_by_report_name(corp_code=cc,bgn_de="20230101",end_de="20260620",
        pblntf_tys="",keywords=[kw],max_pages=4)
    out=[]
    for it in items[:6]:
        html=(await c.get_document_cached(it["rcept_no"])).get("html","") or ""
        flat=re.sub(r"\s+"," ",re.sub(r"<[^>]+>"," ",html))
        amt=_acode_int(html,acode)
        pref=bool(re.search(r"우선주|기타주식|종류주식",flat))
        ds=sum(int(a.replace(",",""))for a in DAILY.findall(flat))
        gap = ds>(amt or 0)*1.05 and pref and ds>0
        out.append({"dt":it["rcept_dt"],"acode":amt,"daily":ds,"pref":pref,"undercount":gap})
    return out
async def one(name):
    try:
        r=await resolve_company_query(name)
        if not r.selected: return {"name":name,"err":"resolve"}
        cc=r.selected["corp_code"]
        acq=await parse_results(cc,"자기주식취득결과","ACQ_AMT")
        dsp=await parse_results(cc,"자기주식처분결과","DSP_AMT")
        return {"name":name,"acq":acq,"dsp":dsp}
    except Exception as e:
        return {"name":name,"err":f"{type(e).__name__}"}
async def main():
    n=int(sys.argv[1]) if len(sys.argv)>1 else 200
    names=ks_top(n)
    print(f"KOSPI 상위 {len(names)}사 자사주 취득/처분결과 복수종류 점검")
    res,t0=[],time.monotonic()
    for i in range(0,len(names),16):
        res.extend(await asyncio.gather(*(one(x) for x in names[i:i+16])))
        d=min(i+16,len(names))
        if d%48==0 or d==len(names): print(f"  ... {d}/{len(names)} ({time.monotonic()-t0:.0f}s)")
    acq_uc=[(r["name"],e) for r in res if r.get("acq") for e in r["acq"] if e["undercount"]]
    dsp_uc=[(r["name"],e) for r in res if r.get("dsp") for e in r["dsp"] if e["undercount"]]
    pref_cos=[r["name"] for r in res if (r.get("acq") or r.get("dsp")) and any(e["pref"] for e in (r.get("acq") or [])+(r.get("dsp") or []))]
    print(f"\n우선주 블록 보유(자사주 활동) 기업: {len(pref_cos)}사")
    print(f"\n=== 취득결과 undercount(ACODE<일별합, 우선주누락): {len(acq_uc)}건 ===")
    for nm,e in acq_uc: print(f"  {nm} {e['dt']}: ACODE {(e['acode'] or 0)/1e8:.0f}억 → 일별 {e['daily']/1e8:.0f}억")
    print(f"\n=== 처분결과 undercount: {len(dsp_uc)}건 ===")
    for nm,e in dsp_uc: print(f"  {nm} {e['dt']}: ACODE {(e['acode'] or 0)/1e8:.0f}억 → 일별 {e['daily']/1e8:.0f}억")
    json.dump(res,open("wiki/architecture/audits/data/260617_treasury_multitype_census.json","w",encoding="utf-8"),ensure_ascii=False,indent=2)
    print("\n저장: 260617_treasury_multitype_census.json")
asyncio.run(main())
